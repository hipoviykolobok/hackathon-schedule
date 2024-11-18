import pandas as pd
import json
import os
import re
from openpyxl import load_workbook

# Укажите путь к вашему Excel файлу
file_path = "schedule.xlsx"

# Загрузите рабочую книгу с помощью openpyxl
wb = load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames

# Функция для нормализации названия дня недели
def normalize_day(day_str):
    return day_str.replace(" ", "").upper()

# Словарь для перевода дней недели
days_map = {
    "ПОНЕДЕЛЬНИК": "Понедельник",
    "ВТОРНИК": "Вторник",
    "СРЕДА": "Среда",
    "ЧЕТВЕРГ": "Четверг",
    "ПЯТНИЦА": "Пятница",
    "СУББОТА": "Суббота",
    "ВОСКРЕСЕНЬЕ": "Воскресенье"
}

# Цвета для определения недели
color_to_week = {
    'FF00FF00': 'Нечетная неделя',  # Зеленый
    'FFFFFF00': 'Четная неделя',  # Желтый
    'FFFF0000': 'Каждая неделя'  # Красный
}

# Заполняем пустые значения в первом столбце (объединённые ячейки)
schedule = []

# Перебор всех листов
for sheet_name in sheet_names:
    print(f"Обработка листа: {sheet_name}")

    # Чтение текущего листа с помощью pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Извлечение названия группы из ячейки C1
    group_name = str(df.iloc[0, 2]).strip()

    # Доступ к листу с помощью openpyxl
    sheet = wb[sheet_name]

    current_day = None
    current_time = None

    # Перебор строк таблицы
    for index, row in df.iterrows():
        # Проверяем и нормализуем день недели
        day_raw = str(row[0]).strip()
        normalized_day = normalize_day(day_raw)
        if normalized_day in days_map:
            current_day = days_map[normalized_day]

        # Проверяем время занятия
        if pd.notna(row[1]) and isinstance(row[1], str) and "Часы" not in row[1]:
            current_time = str(row[1]).replace("\n", "").strip()

        # Проверяем наличие информации о дисциплине
        cell = str(row[2]).strip()
        if pd.notna(row[2]):
            # Пропускаем строки, не содержащие корректных данных
            if not current_day or not current_time:
                continue

            # Получаем цвет ячейки
            cell_color = sheet.cell(row=index + 1, column=3).fill.start_color.index

            # Определяем неделю на основе цвета
            week_info = color_to_week.get(cell_color, 'Неизвестно')

            # Разделяем строку на предмет, преподавателя и аудиторию
            lesson_type = None
            subject = None
            teacher = "НУ"
            room = None

            # Определяем тип занятия
            if cell != "nan":
                if cell.startswith("пр."):
                    lesson_type = "Практика"
                    class_string = cell[3:]  # Удаляем префикс 'пр.'
                elif cell.startswith("лек."):
                    lesson_type = "Лекция"
                    class_string = cell[4:]  # Удаляем префикс 'лек.'
                elif cell.startswith("лаб."):
                    lesson_type = "Лабораторная"
                    class_string = cell[4:]  # Удаляем префикс 'лаб.'

                # Ищем аудиторию
                room_match = re.search(r'\d{2}-\d{3}[аб]?$', class_string) # ищим через регулярное выржаение аудиторию
                if room_match:
                    room = room_match.group()
                    class_string = class_string[:room_match.start()].strip()  # Убираем аудиторию из строки

                # Ищем преподавателя
                teacher_match = re.search(r'[А-ЯЁ]+\s[А-ЯЁ]\.[А-ЯЁ]\.', class_string)
                if teacher_match:
                    teacher = teacher_match.group().strip().upper()  # Преподаватель с приведением к верхнему регистру
                    teacher_parts = teacher.split()  # Разделяем на фамилию и инициалы
                    if teacher_parts:
                        teacher_parts[0] = teacher_parts[
                            0].capitalize()  # Приводим фамилию к нормальному виду (первая буква заглавная)
                        teacher = " ".join(teacher_parts)  # Собираем обратно
                    subject = class_string[:teacher_match.start()].strip()  # Название курса до преподавателя
                else:
                    subject = class_string.strip()  # Если преподаватель не найден, все оставшееся — это название курса

                # Если указана "КпоВ: Базовая физическая культура", преподаватель и аудитория отсутствуют
                if "КпоВ: Базовая физическая культура" in subject:
                    subject = "КпоВ: Базовая физическая культура / Базовые виды спорта"
                    teacher = "CК"
                    room = "СК"

                # Добавляем запись в расписание
                schedule.append({
                    "group": group_name,
                    "day": current_day,
                    "time": current_time,
                    "course": {
                        "name": subject,
                        "type": lesson_type,
                        "teacher": teacher if teacher else None,
                        "room": room if room else None,
                        "week": week_info  # Информация о неделе
                    }
                })

# Сохранение результата в JSON
output_file = "schedule.json"


with open("schedule.json", "w", encoding="utf-8") as f:
    json.dump(schedule, f, ensure_ascii=False, indent=4)

print(f"Расписание сохранено в файл.")
